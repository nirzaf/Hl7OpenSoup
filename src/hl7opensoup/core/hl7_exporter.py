"""
HL7 Data Transformation and Export for HL7 OpenSoup.

This module provides functionality to export HL7 messages to various formats
including CSV, Excel, JSON, XML, and custom formats.
"""

import json
import csv
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
import xml.etree.ElementTree as ET

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from hl7opensoup.models.hl7_message import HL7Message, HL7MessageCollection


class HL7ExportFormat:
    """Export format definitions."""
    
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    XML = "xml"
    HTML = "html"
    TEXT = "text"
    PIPE_DELIMITED = "pipe"


class HL7Exporter:
    """HL7 message exporter with multiple format support."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.logger = logging.getLogger(__name__)
    
    def get_supported_formats(self) -> Dict[str, str]:
        """Get supported export formats.
        
        Returns:
            Dictionary of format_key -> description
        """
        formats = {
            HL7ExportFormat.CSV: "Comma Separated Values (*.csv)",
            HL7ExportFormat.JSON: "JavaScript Object Notation (*.json)",
            HL7ExportFormat.XML: "Extensible Markup Language (*.xml)",
            HL7ExportFormat.HTML: "HyperText Markup Language (*.html)",
            HL7ExportFormat.TEXT: "Plain Text (*.txt)",
            HL7ExportFormat.PIPE_DELIMITED: "Pipe Delimited (*.txt)"
        }
        
        if HAS_OPENPYXL:
            formats[HL7ExportFormat.EXCEL] = "Microsoft Excel (*.xlsx)"
        
        return formats
    
    def export_messages(self, messages: List[HL7Message], file_path: str, 
                       export_format: str, options: Dict[str, Any] = None) -> bool:
        """Export messages to specified format.
        
        Args:
            messages: List of HL7 messages to export
            file_path: Output file path
            export_format: Export format
            options: Export options
            
        Returns:
            True if successful, False otherwise
        """
        if not messages:
            self.logger.error("No messages to export")
            return False
        
        options = options or {}
        
        try:
            if export_format == HL7ExportFormat.CSV:
                return self._export_to_csv(messages, file_path, options)
            elif export_format == HL7ExportFormat.EXCEL:
                return self._export_to_excel(messages, file_path, options)
            elif export_format == HL7ExportFormat.JSON:
                return self._export_to_json(messages, file_path, options)
            elif export_format == HL7ExportFormat.XML:
                return self._export_to_xml(messages, file_path, options)
            elif export_format == HL7ExportFormat.HTML:
                return self._export_to_html(messages, file_path, options)
            elif export_format == HL7ExportFormat.TEXT:
                return self._export_to_text(messages, file_path, options)
            elif export_format == HL7ExportFormat.PIPE_DELIMITED:
                return self._export_to_pipe_delimited(messages, file_path, options)
            else:
                self.logger.error(f"Unsupported export format: {export_format}")
                return False
                
        except Exception as e:
            self.logger.error(f"Export failed: {e}")
            return False
    
    def _export_to_csv(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to CSV format."""
        include_segments = options.get('include_segments', True)
        include_fields = options.get('include_fields', True)
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if include_segments and include_fields:
                # Detailed export with all segments and fields
                writer = csv.writer(csvfile)
                
                # Header
                writer.writerow([
                    'Message_Index', 'Message_Type', 'Control_ID', 'Version',
                    'Segment_Index', 'Segment_Name', 'Field_Index', 'Field_Name', 'Field_Value'
                ])
                
                # Data
                for msg_idx, message in enumerate(messages):
                    msg_type = message.get_message_type()
                    control_id = message.get_control_id()
                    version = message.get_version().value if message.get_version() else ""
                    
                    for seg_idx, segment in enumerate(message.segments):
                        for field_idx, field in enumerate(segment.fields):
                            writer.writerow([
                                msg_idx + 1, msg_type, control_id, version,
                                seg_idx + 1, segment.name, field_idx + 1,
                                f"{segment.name}.{field_idx + 1}", field.get_value()
                            ])
            else:
                # Summary export
                writer = csv.writer(csvfile)
                
                # Header
                writer.writerow([
                    'Message_Index', 'Message_Type', 'Control_ID', 'Version',
                    'Segment_Count', 'Validation_Status', 'Content'
                ])
                
                # Data
                for msg_idx, message in enumerate(messages):
                    status = "Error" if message.has_errors() else "Warning" if message.has_warnings() else "Valid"
                    
                    writer.writerow([
                        msg_idx + 1,
                        message.get_message_type(),
                        message.get_control_id(),
                        message.get_version().value if message.get_version() else "",
                        len(message.segments),
                        status,
                        str(message).replace('\r', '\\r').replace('\n', '\\n')
                    ])
        
        return True
    
    def _export_to_excel(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to Excel format."""
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl not available for Excel export")
            return False
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Message Summary")
        self._create_excel_summary_sheet(summary_ws, messages)
        
        # Create detailed sheets for each message type
        message_types = {}
        for message in messages:
            msg_type = message.get_message_type()
            if msg_type not in message_types:
                message_types[msg_type] = []
            message_types[msg_type].append(message)
        
        for msg_type, type_messages in message_types.items():
            # Clean sheet name
            sheet_name = msg_type.replace('^', '_').replace('|', '_')[:31]
            detail_ws = wb.create_sheet(sheet_name)
            self._create_excel_detail_sheet(detail_ws, type_messages, msg_type)
        
        wb.save(file_path)
        return True
    
    def _create_excel_summary_sheet(self, worksheet, messages: List[HL7Message]):
        """Create Excel summary sheet."""
        # Headers
        headers = [
            'Message #', 'Type', 'Control ID', 'Version', 'Segments',
            'Validation', 'Errors', 'Warnings', 'Timestamp'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, message in enumerate(messages, 2):
            error_count = len([r for r in message.validation_results if r.level.value == "error"])
            warning_count = len([r for r in message.validation_results if r.level.value == "warning"])
            status = "Error" if error_count > 0 else "Warning" if warning_count > 0 else "Valid"
            
            data = [
                row - 1,
                message.get_message_type(),
                message.get_control_id(),
                message.get_version().value if message.get_version() else "",
                len(message.segments),
                status,
                error_count,
                warning_count,
                message.timestamp.strftime("%Y-%m-%d %H:%M:%S") if message.timestamp else ""
            ]
            
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_excel_detail_sheet(self, worksheet, messages: List[HL7Message], message_type: str):
        """Create Excel detail sheet for specific message type."""
        # Headers
        headers = [
            'Message #', 'Segment #', 'Segment', 'Field #', 'Field Value'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        row = 2
        for msg_idx, message in enumerate(messages):
            for seg_idx, segment in enumerate(message.segments):
                for field_idx, field in enumerate(segment.fields):
                    data = [
                        msg_idx + 1,
                        seg_idx + 1,
                        segment.name,
                        field_idx + 1,
                        field.get_value()
                    ]
                    
                    for col, value in enumerate(data, 1):
                        worksheet.cell(row=row, column=col, value=value)
                    row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _export_to_json(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to JSON format."""
        include_raw = options.get('include_raw', False)
        pretty_print = options.get('pretty_print', True)
        
        data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'message_count': len(messages),
                'format': 'HL7 OpenSoup JSON Export'
            },
            'messages': []
        }
        
        for msg_idx, message in enumerate(messages):
            msg_data = {
                'index': msg_idx + 1,
                'message_type': message.get_message_type(),
                'control_id': message.get_control_id(),
                'version': message.get_version().value if message.get_version() else None,
                'timestamp': message.timestamp.isoformat() if message.timestamp else None,
                'segments': []
            }
            
            if include_raw:
                msg_data['raw_content'] = str(message)
            
            # Add validation results
            if message.validation_results:
                msg_data['validation'] = [
                    {
                        'level': result.level.value,
                        'message': result.message,
                        'location': result.location
                    }
                    for result in message.validation_results
                ]
            
            # Add segments
            for seg_idx, segment in enumerate(message.segments):
                seg_data = {
                    'index': seg_idx + 1,
                    'name': segment.name,
                    'fields': []
                }
                
                for field_idx, field in enumerate(segment.fields):
                    field_data = {
                        'index': field_idx + 1,
                        'value': field.get_value(),
                        'components': []
                    }
                    
                    # Add components if they exist
                    for comp_idx, component in enumerate(field.components):
                        comp_data = {
                            'index': comp_idx + 1,
                            'subcomponents': [str(sub) for sub in component.subcomponents]
                        }
                        field_data['components'].append(comp_data)
                    
                    seg_data['fields'].append(field_data)
                
                msg_data['segments'].append(seg_data)
            
            data['messages'].append(msg_data)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            if pretty_print:
                json.dump(data, f, indent=2, ensure_ascii=False)
            else:
                json.dump(data, f, ensure_ascii=False)
        
        return True
    
    def _export_to_xml(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to XML format."""
        root = ET.Element("HL7Messages")
        root.set("exportTimestamp", datetime.now().isoformat())
        root.set("messageCount", str(len(messages)))
        
        for msg_idx, message in enumerate(messages):
            msg_elem = ET.SubElement(root, "Message")
            msg_elem.set("index", str(msg_idx + 1))
            msg_elem.set("type", message.get_message_type())
            msg_elem.set("controlId", message.get_control_id())
            
            if message.get_version():
                msg_elem.set("version", message.get_version().value)
            
            # Add segments
            for seg_idx, segment in enumerate(message.segments):
                seg_elem = ET.SubElement(msg_elem, "Segment")
                seg_elem.set("index", str(seg_idx + 1))
                seg_elem.set("name", segment.name)
                
                # Add fields
                for field_idx, field in enumerate(segment.fields):
                    field_elem = ET.SubElement(seg_elem, "Field")
                    field_elem.set("index", str(field_idx + 1))
                    field_elem.text = field.get_value()
        
        # Write to file
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ", level=0)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        return True
    
    def _export_to_html(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to HTML format."""
        html_content = self._generate_html_report(messages, options)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return True
    
    def _export_to_text(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to plain text format."""
        include_headers = options.get('include_headers', True)
        separator = options.get('separator', '\n' + '='*80 + '\n')
        
        with open(file_path, 'w', encoding='utf-8') as f:
            if include_headers:
                f.write(f"HL7 OpenSoup Export\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Messages: {len(messages)}\n")
                f.write('='*80 + '\n\n')
            
            for msg_idx, message in enumerate(messages):
                if include_headers:
                    f.write(f"Message {msg_idx + 1}:\n")
                    f.write(f"Type: {message.get_message_type()}\n")
                    f.write(f"Control ID: {message.get_control_id()}\n")
                    f.write(f"Version: {message.get_version().value if message.get_version() else 'Unknown'}\n")
                    f.write('-'*40 + '\n')
                
                f.write(str(message))
                
                if msg_idx < len(messages) - 1:
                    f.write(separator)
        
        return True
    
    def _export_to_pipe_delimited(self, messages: List[HL7Message], file_path: str, options: Dict[str, Any]) -> bool:
        """Export messages to pipe-delimited format (standard HL7)."""
        with open(file_path, 'w', encoding='utf-8') as f:
            for msg_idx, message in enumerate(messages):
                f.write(str(message))
                if msg_idx < len(messages) - 1:
                    f.write('\n\n')
        
        return True
    
    def _generate_html_report(self, messages: List[HL7Message], options: Dict[str, Any]) -> str:
        """Generate HTML report for messages."""
        html = """<!DOCTYPE html>
<html>
<head>
    <title>HL7 OpenSoup Export Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .header { background-color: #f0f0f0; padding: 10px; border: 1px solid #ccc; }
        .message { margin: 20px 0; border: 1px solid #ddd; }
        .message-header { background-color: #e8e8e8; padding: 8px; font-weight: bold; }
        .segment { margin: 5px 0; }
        .segment-name { font-weight: bold; color: #0066cc; }
        .field { margin-left: 20px; font-family: monospace; }
        .validation-error { color: red; }
        .validation-warning { color: orange; }
        table { border-collapse: collapse; width: 100%; margin: 10px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
    </style>
</head>
<body>
"""
        
        # Header
        html += f"""
    <div class="header">
        <h1>HL7 OpenSoup Export Report</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Total Messages: {len(messages)}</p>
    </div>
"""
        
        # Summary table
        html += """
    <h2>Message Summary</h2>
    <table>
        <tr>
            <th>Message #</th>
            <th>Type</th>
            <th>Control ID</th>
            <th>Version</th>
            <th>Segments</th>
            <th>Status</th>
        </tr>
"""
        
        for msg_idx, message in enumerate(messages):
            status = "Error" if message.has_errors() else "Warning" if message.has_warnings() else "Valid"
            status_class = "validation-error" if message.has_errors() else "validation-warning" if message.has_warnings() else ""
            
            html += f"""
        <tr>
            <td>{msg_idx + 1}</td>
            <td>{message.get_message_type()}</td>
            <td>{message.get_control_id()}</td>
            <td>{message.get_version().value if message.get_version() else ''}</td>
            <td>{len(message.segments)}</td>
            <td class="{status_class}">{status}</td>
        </tr>
"""
        
        html += """
    </table>
"""
        
        # Detailed messages
        html += "<h2>Message Details</h2>"
        
        for msg_idx, message in enumerate(messages):
            html += f"""
    <div class="message">
        <div class="message-header">
            Message {msg_idx + 1}: {message.get_message_type()} ({message.get_control_id()})
        </div>
        <pre style="background-color: #f8f8f8; padding: 10px; overflow-x: auto;">{str(message)}</pre>
    </div>
"""
        
        html += """
</body>
</html>
"""
        
        return html
